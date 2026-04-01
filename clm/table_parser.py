"""
Table Parser — converts spreadsheet / CSV / TSV files into a list of row-dicts.

Supported formats:
  • CSV / TSV  (stdlib csv module)
  • XLSX       (openpyxl)
  • XLS        (xlrd — optional fallback, or openpyxl with .xlsx re-save)
  • ODS        (openpyxl can read if odfpy installed, else manual XML parse)
  • Google Sheets URL  (public sheets via CSV export URL)

Each row becomes a dict: { col_name: cell_value, ... }
Column headers are sanitised to snake_case identifiers (e.g. "Invoice No." → "invoice_no").

For scanned PDF / image tables, we use Gemini Vision to recreate structured
table data from the visual layout (OCR + AI reconstruction).
"""

import csv
import io
import json
import logging
import re
import hashlib
from pathlib import Path

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Header sanitisation
# ---------------------------------------------------------------------------

def _sanitise_header(raw: str) -> str:
    """
    Convert a raw column header into a clean snake_case identifier.
    "Invoice No."  → "invoice_no"
    "Total Amount (USD)" → "total_amount_usd"
    "  First Name  " → "first_name"
    """
    s = str(raw).strip()
    # Remove chars that aren't alphanumeric, space, or underscore
    s = re.sub(r'[^a-zA-Z0-9\s_]', ' ', s)
    # Collapse whitespace → single underscore
    s = re.sub(r'\s+', '_', s.strip())
    s = s.lower().strip('_')
    return s or 'column'


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Ensure unique headers by appending _2, _3, etc."""
    seen = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            result.append(h)
    return result


def _clean_value(val):
    """Convert cell value to a JSON-friendly type."""
    if val is None:
        return ''
    if isinstance(val, (int, float, bool)):
        return val
    s = str(val).strip()
    # Try numeric conversion
    try:
        if '.' in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        pass
    return s


# ---------------------------------------------------------------------------
# CSV / TSV parser
# ---------------------------------------------------------------------------

def parse_csv(file_bytes: bytes, delimiter: str = ',', encoding: str = 'utf-8') -> dict:
    """
    Parse CSV/TSV bytes → { headers: [...], rows: [{...}, ...], row_count, col_count }.
    Auto-detects delimiter if not specified.
    """
    try:
        text = file_bytes.decode(encoding)
    except UnicodeDecodeError:
        text = file_bytes.decode('latin-1')

    # Auto-detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096])
        delimiter = dialect.delimiter
    except csv.Error:
        pass  # use provided delimiter

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    raw_rows = list(reader)

    if not raw_rows:
        return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

    # First non-empty row = headers
    raw_headers = raw_rows[0]
    headers = _dedupe_headers([_sanitise_header(h) for h in raw_headers])
    col_count = len(headers)

    rows = []
    for raw_row in raw_rows[1:]:
        if not any(cell.strip() for cell in raw_row):
            continue  # skip empty rows
        row = {}
        for i, h in enumerate(headers):
            row[h] = _clean_value(raw_row[i] if i < len(raw_row) else '')
        rows.append(row)

    return {
        'headers': headers,
        'rows': rows,
        'row_count': len(rows),
        'col_count': col_count,
        'original_headers': [str(h).strip() for h in raw_headers],
    }


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

def parse_xlsx(file_bytes: bytes, sheet_name: str | None = None) -> dict:
    """
    Parse XLSX bytes → { headers, rows, row_count, col_count, sheet_names }.
    Requires openpyxl.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            'openpyxl is required for XLSX parsing. '
            'Install it with: pip install openpyxl'
        )

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if not all_rows:
        return {
            'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0,
            'sheet_names': sheet_names,
        }

    # Find first non-empty row as header
    header_idx = 0
    for i, row in enumerate(all_rows):
        if any(cell is not None and str(cell).strip() for cell in row):
            header_idx = i
            break

    raw_headers = all_rows[header_idx]
    # Trim trailing None columns
    while raw_headers and raw_headers[-1] is None:
        raw_headers.pop()

    headers = _dedupe_headers([_sanitise_header(h or f'column_{i+1}') for i, h in enumerate(raw_headers)])
    col_count = len(headers)

    rows = []
    for raw_row in all_rows[header_idx + 1:]:
        # Skip completely empty rows
        if not any(cell is not None and str(cell).strip() for cell in raw_row[:col_count]):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = _clean_value(raw_row[i] if i < len(raw_row) else None)
        rows.append(row)

    return {
        'headers': headers,
        'rows': rows,
        'row_count': len(rows),
        'col_count': col_count,
        'sheet_names': sheet_names,
        'active_sheet': ws.title,
        'original_headers': [str(h).strip() if h else '' for h in all_rows[header_idx][:col_count]],
    }


# ---------------------------------------------------------------------------
# XLS parser (legacy .xls format)
# ---------------------------------------------------------------------------

def parse_xls(file_bytes: bytes, sheet_name: str | None = None) -> dict:
    """
    Parse legacy XLS bytes. Tries xlrd first, falls back to converting
    via openpyxl if xlrd isn't installed.
    """
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheet_names = wb.sheet_names()
        ws = wb.sheet_by_name(sheet_name) if sheet_name and sheet_name in sheet_names else wb.sheet_by_index(0)

        all_rows = []
        for rx in range(ws.nrows):
            all_rows.append([ws.cell_value(rx, cx) for cx in range(ws.ncols)])

        if not all_rows:
            return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0, 'sheet_names': sheet_names}

        raw_headers = all_rows[0]
        headers = _dedupe_headers([_sanitise_header(h or f'column_{i+1}') for i, h in enumerate(raw_headers)])
        col_count = len(headers)

        rows = []
        for raw_row in all_rows[1:]:
            if not any(str(cell).strip() for cell in raw_row[:col_count]):
                continue
            row = {}
            for i, h in enumerate(headers):
                row[h] = _clean_value(raw_row[i] if i < len(raw_row) else '')
            rows.append(row)

        return {
            'headers': headers, 'rows': rows, 'row_count': len(rows),
            'col_count': col_count, 'sheet_names': sheet_names,
            'original_headers': [str(h).strip() for h in all_rows[0]],
        }
    except ImportError:
        # Fallback: try openpyxl (it can sometimes read .xls if saved as xlsx internally)
        logger.warning('xlrd not installed, attempting openpyxl fallback for .xls')
        return parse_xlsx(file_bytes, sheet_name)


# ---------------------------------------------------------------------------
# Google Sheets URL parser
# ---------------------------------------------------------------------------

def parse_google_sheet(url: str) -> dict:
    """
    Parse a public Google Sheet by converting its URL to CSV export URL.
    Supports:
      https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit...
      https://docs.google.com/spreadsheets/d/{SHEET_ID}/
    Optional: ?gid=123 or #gid=123 for specific sheet tab.
    """
    import urllib.request

    # Extract sheet ID
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not match:
        raise ValueError(f'Invalid Google Sheets URL: {url}')
    sheet_id = match.group(1)

    # Extract gid (tab)
    gid_match = re.search(r'[?&#]gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else '0'

    csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'

    try:
        req = urllib.request.Request(csv_url, headers={'User-Agent': 'Drafter-CLM/1.0'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            csv_bytes = resp.read()
    except Exception as e:
        raise ValueError(
            f'Failed to fetch Google Sheet. Make sure it is shared as '
            f'"Anyone with the link can view". Error: {e}'
        )

    result = parse_csv(csv_bytes)
    result['source'] = 'google_sheets'
    result['sheet_id'] = sheet_id
    result['gid'] = gid
    return result


# ---------------------------------------------------------------------------
# ODS parser (OpenDocument Spreadsheet)
# ---------------------------------------------------------------------------

def parse_ods(file_bytes: bytes) -> dict:
    """
    Parse ODS (OpenDocument Spreadsheet) files.
    Uses openpyxl if odfpy is installed, otherwise parses XML directly.
    """
    # Try openpyxl first (requires odfpy for ODS support)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        if not all_rows:
            return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

        raw_headers = all_rows[0]
        while raw_headers and raw_headers[-1] is None:
            raw_headers.pop()
        headers = _dedupe_headers([_sanitise_header(h or f'column_{i+1}') for i, h in enumerate(raw_headers)])
        col_count = len(headers)

        rows = []
        for raw_row in all_rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in raw_row[:col_count]):
                continue
            row = {}
            for i, h in enumerate(headers):
                row[h] = _clean_value(raw_row[i] if i < len(raw_row) else None)
            rows.append(row)

        return {
            'headers': headers, 'rows': rows,
            'row_count': len(rows), 'col_count': col_count,
            'original_headers': [str(h).strip() if h else '' for h in all_rows[0][:col_count]],
        }
    except Exception as e:
        logger.warning(f'openpyxl ODS read failed: {e}, falling back to XML parse')

    # Fallback: manual XML parse for simple ODS
    import zipfile
    import xml.etree.ElementTree as ET

    NS = {
        'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
        'table':  'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
        'text':   'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
    }

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        content_xml = zf.read('content.xml')

    root = ET.fromstring(content_xml)
    body = root.find('.//office:body/office:spreadsheet', NS)
    if body is None:
        return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

    table = body.find('table:table', NS)
    if table is None:
        return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

    all_rows = []
    for tr in table.findall('table:table-row', NS):
        cells = []
        for tc in tr.findall('table:table-cell', NS):
            texts = [t.text or '' for t in tc.findall('.//text:p', NS)]
            cells.append(' '.join(texts).strip())
        all_rows.append(cells)

    if not all_rows:
        return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

    raw_headers = all_rows[0]
    headers = _dedupe_headers([_sanitise_header(h or f'column_{i+1}') for i, h in enumerate(raw_headers)])
    col_count = len(headers)

    rows = []
    for raw_row in all_rows[1:]:
        if not any(str(cell).strip() for cell in raw_row[:col_count]):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = _clean_value(raw_row[i] if i < len(raw_row) else '')
        rows.append(row)

    return {
        'headers': headers, 'rows': rows,
        'row_count': len(rows), 'col_count': col_count,
        'original_headers': [str(h).strip() for h in all_rows[0][:col_count]],
    }


# ---------------------------------------------------------------------------
# AI Table Recreation from PDF / Image (OCR + Gemini Vision)
# ---------------------------------------------------------------------------

def extract_table_from_image_ai(file_bytes: bytes, file_type: str = 'pdf',
                                 api_key: str | None = None) -> dict:
    """
    Use Gemini Vision to extract structured table data from a scanned
    PDF or image file.  Returns the same format as other parsers:
    { headers, rows, row_count, col_count }.

    This handles:
      • Scanned PDFs with tabular data
      • Photos / screenshots of tables
      • Any image format with table content
    """
    import google.generativeai as genai
    import os

    key = api_key or os.environ.get('GEMINI_API_KEY', '')
    if not key:
        raise ValueError('GEMINI_API_KEY is required for AI table extraction')

    genai.configure(api_key=key)

    # Convert PDF first page to image if needed
    mime_type = {
        'pdf':  'application/pdf',
        'png':  'image/png',
        'jpg':  'image/jpeg',
        'jpeg': 'image/jpeg',
        'gif':  'image/gif',
        'bmp':  'image/bmp',
        'tiff': 'image/tiff',
        'tif':  'image/tiff',
        'webp': 'image/webp',
    }.get(file_type.lower(), 'application/octet-stream')

    prompt = """You are a precise data extraction tool. Analyze this document/image and extract ALL tabular data you can find.

Instructions:
1. Identify every table in the document
2. For each table, extract the column headers exactly as they appear
3. Extract every row of data
4. If there are multiple tables, merge them if they have the same columns, or use the largest table
5. Return ONLY valid JSON in this exact format:

{
  "headers": ["Column 1", "Column 2", "Column 3"],
  "rows": [
    {"Column 1": "value1", "Column 2": "value2", "Column 3": "value3"},
    {"Column 1": "value4", "Column 2": "value5", "Column 3": "value6"}
  ]
}

Rules:
- Use the exact column header text from the document
- Preserve all data values exactly (numbers, dates, text)
- Include ALL rows, do not truncate
- If a cell is empty, use empty string ""
- Numbers should be numbers (not strings) when clearly numeric
- Return ONLY the JSON object, no markdown, no explanation"""

    model = genai.GenerativeModel('gemini-2.5-flash')

    response = model.generate_content(
        [
            prompt,
            {'mime_type': mime_type, 'data': file_bytes},
        ],
        generation_config={
            'temperature': 0.1,
            'max_output_tokens': 8192,
        },
    )

    # Parse the JSON response
    text = response.text.strip()
    # Strip markdown code fences if present
    if text.startswith('```'):
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text)

    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        logger.error(f'AI table extraction returned invalid JSON: {text[:500]}')
        raise ValueError(f'AI could not parse table from this file: {e}')

    raw_headers = data.get('headers', [])
    raw_rows = data.get('rows', [])

    headers = _dedupe_headers([_sanitise_header(h) for h in raw_headers])

    # Normalise rows to use sanitised headers
    rows = []
    for raw_row in raw_rows:
        row = {}
        if isinstance(raw_row, dict):
            for orig_h, san_h in zip(raw_headers, headers):
                row[san_h] = _clean_value(raw_row.get(orig_h, ''))
        elif isinstance(raw_row, list):
            for i, h in enumerate(headers):
                row[h] = _clean_value(raw_row[i] if i < len(raw_row) else '')
        rows.append(row)

    return {
        'headers': headers,
        'rows': rows,
        'row_count': len(rows),
        'col_count': len(headers),
        'original_headers': [str(h).strip() for h in raw_headers],
        'source': 'ai_vision',
        'ai_model': 'gemini-2.5-flash',
    }


# ---------------------------------------------------------------------------
# Multi-page PDF table extraction (page-by-page AI vision)
# ---------------------------------------------------------------------------

def extract_tables_from_pdf_ai(file_bytes: bytes,
                                api_key: str | None = None,
                                max_pages: int = 50) -> dict:
    """
    Extract tables from a multi-page PDF by rendering each page to an image
    and running Gemini Vision on each. Merges tables with compatible headers.
    Falls back to single-call extraction for short PDFs.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        # Fallback: send whole PDF to Gemini (works for short PDFs)
        return extract_table_from_image_ai(file_bytes, 'pdf', api_key)

    doc = fitz.open(stream=file_bytes, filetype='pdf')
    page_count = min(len(doc), max_pages)

    if page_count <= 3:
        # For short PDFs, send the whole file directly
        doc.close()
        return extract_table_from_image_ai(file_bytes, 'pdf', api_key)

    # For longer PDFs, process page by page and merge
    all_rows = []
    merged_headers = None
    original_headers = None

    for page_num in range(page_count):
        page = doc[page_num]
        # Render page to PNG at 200 DPI for good OCR quality
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes('png')

        try:
            result = extract_table_from_image_ai(img_bytes, 'png', api_key)
            if result['row_count'] > 0:
                if merged_headers is None:
                    merged_headers = result['headers']
                    original_headers = result.get('original_headers', result['headers'])
                    all_rows.extend(result['rows'])
                elif result['headers'] == merged_headers:
                    # Same table continuing on next page
                    all_rows.extend(result['rows'])
                else:
                    # Different table on this page — still add rows, remap columns
                    for row in result['rows']:
                        mapped = {}
                        for h in merged_headers:
                            mapped[h] = row.get(h, '')
                        all_rows.append(mapped)
        except Exception as e:
            logger.warning(f'AI table extraction failed for page {page_num + 1}: {e}')
            continue

    doc.close()

    if merged_headers is None:
        return {'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0}

    return {
        'headers': merged_headers,
        'rows': all_rows,
        'row_count': len(all_rows),
        'col_count': len(merged_headers),
        'original_headers': original_headers,
        'source': 'ai_vision_multipage',
        'ai_model': 'gemini-2.5-flash',
        'pages_processed': page_count,
    }


# ---------------------------------------------------------------------------
# Unified dispatcher
# ---------------------------------------------------------------------------

TABLE_EXTENSIONS = {
    'csv', 'tsv', 'xlsx', 'xls', 'ods',
    # Image/PDF for AI table extraction
    'pdf', 'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'tif', 'webp',
}

def parse_table_file(file_bytes: bytes, filename: str,
                     sheet_name: str | None = None,
                     ai_extract: bool = True,
                     google_sheet_url: str | None = None) -> dict:
    """
    Unified entry point: detect file type and parse accordingly.

    Returns: {
        headers: ['col_a', 'col_b', ...],
        rows: [{ col_a: val, col_b: val, ... }, ...],
        row_count: N,
        col_count: M,
        file_hash: 'sha256...',
        parse_method: 'csv' | 'xlsx' | 'xls' | 'ods' | 'google_sheets' | 'ai_vision',
        ...
    }
    """
    # Google Sheets URL
    if google_sheet_url:
        result = parse_google_sheet(google_sheet_url)
        result['parse_method'] = 'google_sheets'
        result['file_hash'] = hashlib.sha256(google_sheet_url.encode()).hexdigest()
        return result

    ext = Path(filename).suffix.lower().lstrip('.')

    # Content hash
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    if ext in ('csv', 'tsv'):
        delimiter = '\t' if ext == 'tsv' else ','
        result = parse_csv(file_bytes, delimiter=delimiter)
        result['parse_method'] = ext
    elif ext == 'xlsx':
        result = parse_xlsx(file_bytes, sheet_name)
        result['parse_method'] = 'xlsx'
    elif ext == 'xls':
        result = parse_xls(file_bytes, sheet_name)
        result['parse_method'] = 'xls'
    elif ext == 'ods':
        result = parse_ods(file_bytes)
        result['parse_method'] = 'ods'
    elif ext == 'pdf' and ai_extract:
        result = extract_tables_from_pdf_ai(file_bytes)
        result['parse_method'] = 'ai_vision'
    elif ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'tif', 'webp') and ai_extract:
        result = extract_table_from_image_ai(file_bytes, ext)
        result['parse_method'] = 'ai_vision'
    else:
        raise ValueError(
            f'Unsupported table file format: .{ext}. '
            f'Supported: CSV, TSV, XLSX, XLS, ODS, or PDF/images with AI extraction.'
        )

    result['file_hash'] = file_hash
    return result


def rows_to_workflow_documents(parsed: dict, workflow, organization,
                                input_node=None, user=None) -> list:
    """
    Convert parsed table rows into WorkflowDocument objects.
    Each row becomes a WorkflowDocument with:
      - title = "Row {n}" (or a meaningful title from a 'name'/'title' column)
      - file_type = 'csv' (virtual — no physical file)
      - extracted_metadata = the row dict (columns become metadata fields)
      - global_metadata = the row dict + _source tag
      - extraction_status = 'completed'
      - original_text = JSON serialisation of the row

    Returns list of created WorkflowDocument objects.
    """
    from .models import WorkflowDocument

    headers = parsed.get('headers', [])
    rows = parsed.get('rows', [])

    # Try to find a "name" or "title" column for document titles
    title_key = None
    for candidate in ['name', 'title', 'document_name', 'document_title',
                       'file_name', 'filename', 'company_name', 'company',
                       'party_name', 'client_name', 'vendor_name',
                       'contract_name', 'description', 'subject', 'id']:
        if candidate in headers:
            title_key = candidate
            break

    created = []
    for idx, row in enumerate(rows, start=1):
        # Determine title
        if title_key and row.get(title_key):
            title = f"{row[title_key]}"
        else:
            title = f"Row {idx}"

        # Row number as extra metadata
        row_meta = {**row, '_row_number': idx, '_source': 'table'}

        # Content hash for dedup
        row_json = json.dumps(row, sort_keys=True, default=str)
        row_hash = hashlib.sha256(row_json.encode()).hexdigest()

        # Check for duplicate by hash
        existing = WorkflowDocument.objects.filter(
            workflow=workflow,
            file_hash=row_hash,
            extraction_status__in=('completed', 'pending', 'processing'),
        ).first()
        if existing:
            created.append(existing)
            continue

        doc = WorkflowDocument(
            workflow=workflow,
            organization=organization,
            input_node=input_node,
            title=title,
            file_type='csv',
            file_size=len(row_json.encode()),
            direct_text=row_json,
            original_text=row_json,
            text_source='direct',
            extracted_metadata=row,
            global_metadata=row_meta,
            overall_confidence=1.0,
            extraction_status='completed',
            uploaded_by=user,
            file_hash=row_hash,
        )
        # Save without file field (virtual document)
        doc.save()
        created.append(doc)

    return created
